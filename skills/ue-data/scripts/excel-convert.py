#!/usr/bin/env python3
"""Convert between Excel XLSX and CSV/JSON for UE DataTable round-trips.

Usage:
    python excel-convert.py <input> <output> [sheet_name]

Supported conversions:
    .xlsx -> .csv    Export spreadsheet to CSV (UTF-8, UE-compatible)
    .csv  -> .xlsx   Import CSV into spreadsheet for editing
    .xlsx -> .json   Export spreadsheet to JSON (UE DataTable format)
    .json -> .xlsx   Import JSON DataTable export into spreadsheet

Requirements:
    pip install openpyxl
"""
import csv
import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print('ERROR: openpyxl is required. Install with: pip install openpyxl')
    sys.exit(1)


def xlsx_to_csv(xlsx_path, csv_path, sheet_name=None):
    """Convert XLSX sheet to CSV (UTF-8, suitable for UE import)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(['' if v is None else v for v in row])
    wb.close()
    print('OK: {} -> {}'.format(xlsx_path, csv_path))


def csv_to_xlsx(csv_path, xlsx_path, sheet_name='DataTable'):
    """Convert CSV to XLSX for editing in Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    wb.save(xlsx_path)
    print('OK: {} -> {}'.format(csv_path, xlsx_path))


def xlsx_to_json(xlsx_path, json_path, sheet_name=None):
    """Convert XLSX to JSON array format for UE DataTable import."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print('ERROR: Empty spreadsheet')
        return
    headers = [str(h) for h in rows[0]]
    # First column is RowName -> maps to "Name" in JSON
    if headers[0] == 'RowName':
        headers[0] = 'Name'
    result = []
    for row in rows[1:]:
        entry = {}
        for i, val in enumerate(row):
            if i < len(headers):
                key = headers[i]
                if val is None:
                    val = ''
                entry[key] = val
        if entry.get('Name', ''):
            result.append(entry)
    wb.close()
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    print('OK: {} -> {} ({} rows)'.format(xlsx_path, json_path, len(result)))


def json_to_xlsx(json_path, xlsx_path, sheet_name='DataTable'):
    """Convert JSON DataTable export to XLSX for editing."""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if not data:
        print('ERROR: Empty JSON')
        return
    # Collect all keys preserving order
    headers = list(dict.fromkeys(k for row in data for k in row.keys()))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Rename "Name" back to "RowName" for UE convention
    display_headers = ['RowName' if h == 'Name' else h for h in headers]
    ws.append(display_headers)
    for row in data:
        ws.append([row.get(h, '') for h in headers])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    wb.save(xlsx_path)
    print('OK: {} -> {} ({} rows)'.format(json_path, xlsx_path, len(data)))


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python excel-convert.py <input> <output> [sheet_name]')
        print('  Supported: .xlsx <-> .csv, .xlsx <-> .json')
        sys.exit(1)

    inp, out = sys.argv[1], sys.argv[2]
    sheet = sys.argv[3] if len(sys.argv) > 3 else None
    ext_in = os.path.splitext(inp)[1].lower()
    ext_out = os.path.splitext(out)[1].lower()

    if ext_in == '.xlsx' and ext_out == '.csv':
        xlsx_to_csv(inp, out, sheet)
    elif ext_in == '.csv' and ext_out == '.xlsx':
        csv_to_xlsx(inp, out, sheet or 'DataTable')
    elif ext_in == '.xlsx' and ext_out == '.json':
        xlsx_to_json(inp, out, sheet)
    elif ext_in == '.json' and ext_out == '.xlsx':
        json_to_xlsx(inp, out, sheet or 'DataTable')
    else:
        print('ERROR: Unsupported conversion: {} -> {}'.format(ext_in, ext_out))
        sys.exit(1)
